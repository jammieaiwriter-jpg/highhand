# -*- coding: utf-8 -*-
"""93H 進度儀表板產生器
讀取資料夾中最新的「93H進度盤點_*.xlsx」，產出「93H儀表板.html」。
每週三改完 Excel 後，雙擊「產出93H儀表板.bat」即可。
"""
import glob
import html
import json
import os
import re
from datetime import date, timedelta

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
OUT_HTML = os.path.join(BASE, "93H儀表板.html")
TODAY = date.today()

# 工地回報 Apps Script 網址（部署後填入 https://script.google.com/macros/s/…/exec）。
# 填了之後儀表板頂部會出現「📥 今日工地回報（即時）」，由瀏覽器端直接抓，
# 不依賴本機重產——下班/週末工地回報也能即時看到。
REPORT_API = "https://script.google.com/macros/s/AKfycbxxBXks2VtVjBJZhkzkzTVexEpUmcyCKMKjE18-dpiEh2SnfHcwP2VG6SMPV1NfAblJjw/exec"

# ---------------- 里程碑 ----------------
# ★ 治理規則：里程碑為長官核定版，只有玲嬅同意才能改（柏銘表動到里程碑時
#   只能「標示警告」，不得逕改此處）。詳見 工地/AGENTS.md。
MILESTONES = [
    ("1F門禁完成(玻璃/新美門/鐵捲門/防火門)", date(2026, 9, 30), ""),
    ("拆圍籬、景觀開挖", date(2026, 10, 1), "借地10/1~116/2/28已定"),
    ("2F~1F外牆拆架完成", date(2026, 11, 15), "9/2會議三度延(11/13~11/15拆);目標提前待2F~7F清潔後調整"),
    ("消防檢查開始", date(2026, 11, 15), "9/2會議延11/15~116/1/15⚠️尾端壓使照;前置:消防變更9/15~10/15"),
    ("無障礙掛件", date(2026, 11, 15), ""),
    ("無障礙、使管處看現場", date(2026, 12, 5), "使照取得前送資料即可"),
    ("取得使用執照", date(2027, 1, 15), ""),
]

# ---------------- 工具 ----------------
DONE_PAT = re.compile(r"已完成|已進場|完成$")
READY_PAT = re.compile(r"已可安裝|可安排")


def minguo(d: date) -> str:
    return f"{d.year - 1911}/{d.month}/{d.day}"


def parse_date(text: str):
    """從 115/09/01、116.03.30、115/8/21前完工、115/9月底 之類字串抓出西元日期。"""
    if not text:
        return None
    m = re.search(r"(11[4-9])[./](\d{1,2})[./](\d{1,2})", text)
    if m:
        y, mo, dd = int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, dd)
        except ValueError:
            return None
    m = re.search(r"(11[4-9])[./](\d{1,2})\s*月(底|初|中)?", text)
    if m:
        y, mo = int(m.group(1)) + 1911, int(m.group(2))
        part = m.group(3) or "底"
        dd = {"初": 5, "中": 15}.get(part, 28)
        try:
            return date(y, mo, dd)
        except ValueError:
            return None
    return None


def is_yellow(cell) -> bool:
    f = cell.fill
    if f is None or f.patternType != "solid":
        return False
    rgb = getattr(f.fgColor, "rgb", None) or ""
    return str(rgb).upper().endswith(("FFFF00", "FFE699", "FFF2CC", "FFEB9C"))


def classify(status_text: str, note: str):
    """回傳 (bucket, 到期日)。bucket: done/overdue/soon/later/ready/waiting"""
    s = (status_text or "").strip()
    n = (note or "").strip()
    if DONE_PAT.search(s):
        return "done", None
    if "施作中" in s or "安裝中" in s:
        d = parse_date(n)
        if d:
            if d < TODAY:
                return "overdue", d
            if d <= TODAY + timedelta(days=14):
                return "soon", d
            return "later", d
        return "doing", None
    d = parse_date(s)
    if d:
        if d < TODAY:
            return "overdue", d
        if d <= TODAY + timedelta(days=14):
            return "soon", d
        return "later", d
    if READY_PAT.search(s) or READY_PAT.search(n):
        return "ready", None
    return "waiting", None


# ---------------- 讀資料源 ----------------
# 正本＝Google Sheet「93H進度盤點雲端」（2026-09-03 雲端化）。每次執行先抓雲端匯出；
# 抓不到（斷網）才退回本機最新 xlsx。本機 93H進度盤點_1150826.xlsx 已封存不再維護。
CLOUD_XLSX = ("https://docs.google.com/spreadsheets/d/"
              "1JBdrqfzQ_AF0fG8x_RT7wjZSTCjCfgdgnX9FRkQJSd0/export?format=xlsx")
CLOUD_CACHE = os.path.join(BASE, "_雲端盤點快取.xlsx")
SRC = None
try:
    import urllib.request
    req = urllib.request.Request(CLOUD_XLSX, headers={"User-Agent": "93H-dashboard"})
    data = urllib.request.urlopen(req, timeout=30).read()
    if len(data) > 10000:
        with open(CLOUD_CACHE, "wb") as _f:
            _f.write(data)
        SRC = CLOUD_CACHE
        print("資料源：雲端 93H進度盤點雲端（Google Sheets）")
except Exception as _e:
    print(f"雲端抓取失敗（{_e}），改用本機檔")
if SRC is None:
    files = sorted(glob.glob(os.path.join(BASE, "93H進度盤點_*.xlsx")))
    if not files:
        raise SystemExit("找不到資料源（雲端失敗且無本機 93H進度盤點_*.xlsx）")
    SRC = files[-1]
SRC_LABEL = "93H進度盤點雲端(Google Sheet)" if SRC == CLOUD_CACHE else os.path.basename(SRC)
wb = load_workbook(SRC, data_only=True)

items = []  # dict: 類別, 位置, 項目, 前段(送樣/丈量), 日期狀態, 備註, 使照前, bucket, due


def add(cat, loc, name, pre, status, note, permit, prog="", ms=""):
    b, due = classify(status, note)
    items.append(dict(cat=cat, loc=loc, name=name, pre=pre, status=status,
                      note=note, permit=permit, bucket=b, due=due, prog=prog,
                      ms=ms or "報完工"))


def txt(v):
    return "" if v is None else str(v).strip()


ws = wb["燈具進度"]
for row in ws.iter_rows(min_row=2):
    vals = [txt(c.value) for c in row[:7]]
    loc, _, name, sample, plan, note = vals[:6]
    if not name:
        continue
    permit = is_yellow(row[2])
    add("燈具", loc, name, sample, plan, note, permit, ms=vals[6] if len(vals) > 6 else "")

ws = wb["石材追蹤"]
for row in ws.iter_rows(min_row=2):
    vals = [txt(c.value) for c in row[:7]]
    grp, name, status, note = vals[0], vals[1], vals[2], vals[3]
    permit = (len(vals) > 4 and vals[4].upper() in ("V", "★")) or is_yellow(row[1])
    if not name:
        continue
    add("石材", grp, name, "", status, note, permit,
        prog=vals[5] if len(vals) > 5 else "", ms=vals[6] if len(vals) > 6 else "")

ws = wb["磁磚未進場"]
for row in ws.iter_rows(min_row=2):
    vals = [txt(c.value) for c in row[:4]]
    name, size, status = vals[:3]
    if not name:
        continue
    add("磁磚", size, name, "", status, "", is_yellow(row[0]), ms=vals[3] if len(vals) > 3 else "")

ws = wb["新美鐵件"]
for row in ws.iter_rows(min_row=2):
    vals = [txt(c.value) for c in row[:10]]
    grp, loc, _, name, measure, install, note = vals[0], vals[1], vals[2], vals[3], vals[4], vals[5], vals[6]
    permit = (len(vals) > 7 and vals[7].upper() in ("V", "★")) or is_yellow(row[3])
    if not name:
        continue
    status = install if (install and not DONE_PAT.search(measure)) or DONE_PAT.search(install) else (install or measure)
    add("鐵件", f"{grp}/{loc}", name, measure, status or install, note, permit,
        prog=vals[8] if len(vals) > 8 else "", ms=vals[9] if len(vals) > 9 else "")

# ---- 出工預核 / 缺失改善 / 粗工打石 ----
def rows_of(name, ncol):
    if name not in wb.sheetnames:
        return []
    out = []
    for row in wb[name].iter_rows(min_row=2, values_only=True):
        vals = [txt(v) for v in (row[:ncol] if row else [])]
        if any(vals):
            out.append(vals + [""] * (ncol - len(vals)))
    return out

dispatch_all = rows_of("出工預核", 6)
def d_key(r):
    d = parse_date(r[0])
    return d or date.max
dispatch_upcoming = sorted(
    [r for r in dispatch_all
     if parse_date(r[0]) == TODAY or ((parse_date(r[0]) or date.max) > TODAY and not r[4])],
    key=lambda r: (d_key(r), r[1]))
dispatch_today = [r for r in dispatch_upcoming if parse_date(r[0]) == TODAY]
dispatch_unverified = [r for r in dispatch_all if (parse_date(r[0]) or date.max) < TODAY and not r[4]]
FAIL_PAT = re.compile(r"未出|未完成|未進|未派|改期")
dispatch_failed = sorted(
    [r for r in dispatch_all
     if (parse_date(r[0]) or date.max) <= TODAY and FAIL_PAT.search(str(r[4]))
     and "已重排" not in str(r[4]) and "已完成" not in str(r[4])],
    key=d_key)

mat_all = rows_of("進料追蹤", 6)
mat_upcoming = sorted(
    [r for r in mat_all
     if parse_date(r[0]) == TODAY or ((parse_date(r[0]) or date.max) > TODAY and not r[4]) or not r[0]],
    key=d_key)
mat_unverified = [r for r in mat_all if r[0] and (parse_date(r[0]) or date.max) < TODAY and not r[4]]

admin_all = rows_of("行政時程", 6)
admin_upcoming = sorted(
    [r for r in admin_all
     if parse_date(r[0]) == TODAY or ((parse_date(r[0]) or date.max) > TODAY and not r[4])],
    key=d_key)
admin_unverified = [r for r in admin_all if (parse_date(r[0]) or date.max) < TODAY and not r[4]]
admin_failed = [r for r in admin_all
                if r[0] and (parse_date(r[0]) or date.max) <= TODAY
                and re.search(r"改期|未", str(r[4])) and "已重排" not in str(r[4])]

decisions = [r for r in rows_of("待裁決", 6) if not r[3]]

defects = [r for r in rows_of("缺失改善", 8) if r[6] not in ("已完成", "結案")]

labor = rows_of("粗工打石", 8)
def month_key(r):
    d = parse_date(r[0])
    return (d.year, d.month) if d else None
this_month = (TODAY.year, TODAY.month)
labor_stats = {}
for r in labor:
    who = r[5] or "未填"
    try:
        n = float(re.sub(r"[^\d.]", "", str(r[3])) or 0)
    except ValueError:
        n = 0
    amt = 0
    try:
        amt = float(re.sub(r"[^\d.]", "", str(r[6])) or 0)
    except ValueError:
        pass
    s = labor_stats.setdefault(who, {"m_n": 0, "m_amt": 0, "t_n": 0, "t_amt": 0})
    s["t_n"] += n; s["t_amt"] += amt
    if month_key(r) == this_month:
        s["m_n"] += n; s["m_amt"] += amt

if "門禁追蹤" in wb.sheetnames:
    ws = wb["門禁追蹤"]
    for row in ws.iter_rows(min_row=2):
        vals = [txt(c.value) for c in row[:7]]
        vendor, name, status, note = vals[0], vals[1], vals[2], vals[3]
        if not name:
            continue
        permit = (len(vals) > 4 and vals[4].upper() in ("V", "★"))
        add("門禁" if (vals[6] if len(vals) > 6 else "") == "1F門禁" else "工項", vendor, name, "", status, note, permit,
            prog=vals[5] if len(vals) > 5 else "", ms=vals[6] if len(vals) > 6 else "1F門禁")

submissions = rows_of("送審變更", 7)
for s in submissions:
    if s[5] in ("1F門禁", "外牆拆架", "消防檢查", "使照檢查") and s[0] != "消防檢查(消檢)":
        st = s[4] and "已核准" or (s[3] or s[2])
        add("送審", s[1], s[0], "", "已完成" if s[4] else (s[3] or s[2]), s[6], True, ms=s[5])

work_rows = []
ws = wb["發包與工作事項"]
for row in ws.iter_rows(min_row=2, values_only=True):
    vals = [txt(v) for v in (row[:4] if row else [])]
    if len(vals) < 2 or not vals[1]:
        continue
    if "完成" == vals[2].strip() or (vals[2].strip().startswith("完成") and len(vals[2].strip()) <= 4):
        continue
    work_rows.append(vals + [""] * (4 - len(vals)))

# ---------------- 統計 ----------------
open_items = [i for i in items if i["bucket"] != "done"]
overdue = sorted([i for i in open_items if i["bucket"] == "overdue"], key=lambda x: x["due"])
soon = sorted([i for i in open_items if i["bucket"] == "soon"], key=lambda x: x["due"])
ready = [i for i in open_items if i["bucket"] == "ready"]
doing = [i for i in open_items if i["bucket"] == "doing"]
later = sorted([i for i in open_items if i["bucket"] == "later"], key=lambda x: x["due"])
waiting = [i for i in open_items if i["bucket"] == "waiting"]
permit_open = [i for i in open_items if i["permit"]]

BUCKET_LABEL = {
    "overdue": "已過預定日", "soon": "兩週內到期", "ready": "已可施作待排",
    "later": "已排程", "waiting": "等前置條件", "done": "已完成", "doing": "施作中",
}
BUCKET_CLASS = {
    "overdue": "b-red", "soon": "b-orange", "ready": "b-blue",
    "later": "b-gray", "waiting": "b-gray", "done": "b-green", "doing": "b-green",
}


def esc(s):
    return html.escape(str(s)).replace("\n", "<br>")


def badge(i):
    lbl = BUCKET_LABEL[i["bucket"]]
    if i["due"]:
        delta = (i["due"] - TODAY).days
        lbl += f"｜{minguo(i['due'])}" + (f"（逾{-delta}天）" if delta < 0 else f"（{delta}天後）")
    return f'<span class="badge {BUCKET_CLASS[i["bucket"]]}">{esc(lbl)}</span>'


def item_rows(lst, show_cat=True):
    out = []
    for i in lst:
        star = "⭐" if i["permit"] else ""
        cat = f'<td class="cat">{esc(i["cat"])}</td>' if show_cat else ""
        note = i["note"] or i["pre"]
        if i.get("prog"):
            note = f"<b>{esc(i['prog'])}</b>" + ("｜" + esc(note) if note else "")
        else:
            note = esc(note)
        out.append(
            f"<tr>{cat}<td>{esc(i['loc'])}</td>"
            f"<td>{star}{esc(i['name'])}</td><td>{badge(i)}</td>"
            f"<td class='note'>{note}</td></tr>")
    return "\n".join(out)


def section_table(title, lst, tid, note=""):
    if not lst:
        return ""
    head = "<tr><th>類別</th><th>位置</th><th>項目</th><th>狀態</th><th>備註/前置</th></tr>"
    return (f'<details class="sec" id="{tid}"><summary>{title}'
            f'<span class="cnt">{len(lst)}</span>{note}</summary>'
            f'<table>{head}{item_rows(lst)}</table></details>')


# ---------------- 里程碑反推樹 ----------------
# (名稱, 顯示日, 說明, 逾期判定日, 預設展開, 顯示風險警告)
MS_ORDER = [
    ("1F門禁", date(2026, 9, 30), "玻璃/新美門/鐵捲門(青岱)/防火門(佳昇)——拆圍籬(10月初)的前置", date(2026, 9, 30), True, True),
    ("拆架2-7F", date(2026, 10, 7), "9/1會議:因雨延兩週,7F~2F拆架約10/7~10/14；標準層外觀與清潔須趕在該區拆架前完成", date(2026, 10, 14), True, True),
    ("拆架1-2F", date(2026, 11, 13), "2F~1F拆架11/13~11/15(9/2會議)：車道格柵/車道天花板油漆/2F~1F清潔(11/5~11/12)須先完成", date(2026, 11, 15), True, True),
    ("消防檢查", date(2026, 11, 15), "消檢11/15~116/1/15(9/2會議延,尾端壓使照⚠️)；前置鏈：電力變更→消防變更→送電/設備安裝", date(2026, 11, 15), True, True),
    ("使照檢查", date(2026, 12, 5), "12月初無障礙+使管處看現場查核的必要工項；含景觀(園藝核對)；無障礙掛件11/15先行", date(2026, 12, 5), True, True),
    ("景觀工程", date(2026, 10, 1), "借地期間10/1~116/2/28施作(暫置土方/回填)；9/7重排景觀進度；園藝核對項在使照檢查", date(2027, 2, 28), False, False),
    ("報完工", date(2026, 12, 30), "總經理指示：使照前完成，不卡使照", date(2026, 12, 30), False, False),
    ("使照後二工", date(2027, 4, 30), "使照後施作：二工件＋公設裝潢＋樣品屋3A1/13A3＋接待中心S1(先建後售,代銷進駐)", date(2027, 4, 30), False, False),
]
GATING_MS = {"1F門禁", "拆架2-7F", "拆架1-2F", "消防檢查", "使照檢查"}
MS_ICON = {"報完工": "📦", "使照後二工": "🔧", "景觀工程": "🌿"}
MS_LABEL = {"報完工": "使照前完整度", "1F門禁": "門禁(1F/MF/2F)"}
SUB_DEF = [
    ("overdue", "🔴 已過期", True),
    ("soon", "🟠 兩週內到期", True),
    ("later", "📅 已安排", True),
    ("ready", "🔵 可施作但未排日期", True),
    ("doing", "🟢 施作中(缺完成日)", True),
    ("waiting", "⏳ 卡住－等前置", True),
]

def tree_html():
    parts = []
    for ms_name, ms_date, ms_note, ms_deadline, ms_open, ms_risk in MS_ORDER:
        group = [i for i in items if i["ms"] == ms_name]
        if not group:
            continue
        done = [i for i in group if i["bucket"] == "done"]
        open_i = [i for i in group if i["bucket"] != "done"]
        pct = int(len(done) / len(group) * 100) if group else 0
        delta = (ms_date - TODAY).days
        risky = [i for i in open_i if i["due"] and i["due"] > ms_deadline] if ms_risk else []
        head = (f'<div class="ms-head">'
                f'<span class="ms-meta">完成 {len(done)}/{len(group)}（{pct}%）</span>'
                f'<div class="bar"><div class="bar-in" style="width:{pct}%"></div></div>'
                f'<div class="ms-sub">{esc(ms_note)}</div>'
                + (f'<div class="ms-risk">⚠️ {len(risky)} 項排程日晚於里程碑：'
                   + "、".join(esc(i["name"]) for i in risky[:6]) + "</div>" if risky else "")
                + "</div>")
        subs = []
        for key, label, _ in SUB_DEF:
            lst = [i for i in open_i if i["bucket"] == key]
            if key == "waiting":
                lst = sorted(lst, key=lambda x: x["note"] or x["pre"])
            elif key in ("overdue", "soon", "later"):
                lst = sorted(lst, key=lambda x: x["due"] or date.max)
            if not lst:
                continue
            subs.append(f'<div class="subh">{label}（{len(lst)}）</div>'
                        f'<table>{item_rows(lst)}</table>')
        cd = f"剩{delta}天" if delta >= 0 else f"逾{-delta}天"
        cd_cls = ' style="color:#c0392b"' if (delta < 21 and ms_risk) else ''
        icon = MS_ICON.get(ms_name, "🎯")
        label = MS_LABEL.get(ms_name, ms_name)
        open_attr = ""
        parts.append(f'<details{open_attr} class="sec ms-tree"><summary>{icon} {esc(label)}'
                     f'<span class="ms-date-tag">{minguo(ms_date)}</span>'
                     f'<span class="ms-cd-tag"{cd_cls}>{cd}</span>'
                     f'<span class="cnt">{len(open_i)}項未完</span>'
                     f'<span class="pct">{pct}%</span></summary>{head}{"".join(subs)}</details>')
    return "\n".join(parts)


def kpi_html():
    def stats(subset):
        o = sum(1 for i in subset if i["bucket"] == "overdue")
        s = sum(1 for i in subset if i["bucket"] == "soon")
        w = sum(1 for i in subset if i["bucket"] == "waiting")
        d = sum(1 for i in subset if i["bucket"] == "done")
        return o, s, w, d, len(subset)
    gating = [i for i in items if i["ms"] in GATING_MS]
    rest = [i for i in items if i["ms"] not in GATING_MS]
    go, gs, gw, gd, gn = stats(gating)
    ro, rs, rw, rd, rn = stats(rest)
    return f"""
<div class="kpi-row-label">🎯 卡里程碑（使照關鍵路徑）</div>
<div class="kpis">
<div class="kpi k-red"><div class="num">{go}</div><div class="lbl">已過期</div></div>
<div class="kpi k-orange"><div class="num">{gs}</div><div class="lbl">兩週內到期</div></div>
<div class="kpi k-star"><div class="num">{gw}</div><div class="lbl">卡前置</div></div>
<div class="kpi k-green"><div class="num">{gd}/{gn}</div><div class="lbl">已完成</div></div>
</div>
<div class="kpi-row-label sub">📦 完整度清單（不卡使照時程）</div>
<div class="kpis kpis-sub">
<div class="kpi k-red"><div class="num">{ro}</div><div class="lbl">已過期</div></div>
<div class="kpi k-orange"><div class="num">{rs}</div><div class="lbl">兩週內到期</div></div>
<div class="kpi k-star"><div class="num">{rw}</div><div class="lbl">卡前置</div></div>
<div class="kpi k-green"><div class="num">{rd}/{rn}</div><div class="lbl">已完成</div></div>
</div>"""


# ---------------- 里程碑 HTML ----------------
ms_html = []
for name, d, note in MILESTONES:
    delta = (d - TODAY).days
    cls = "ms-past" if delta < 0 else ("ms-hot" if delta <= 21 else "ms-ok")
    cd = f"逾{-delta}天" if delta < 0 else f"剩{delta}天"
    ms_html.append(
        f'<div class="ms {cls}"><div class="ms-date">{minguo(d)}</div>'
        f'<div class="ms-name">{esc(name)}</div>'
        f'<div class="ms-cd">{cd}</div>'
        + (f'<div class="ms-note">{esc(note)}</div>' if note else "") + "</div>")

license_day = (MILESTONES[-1][1] - TODAY).days

work_93h = [r for r in work_rows if r[0] in ("海興段", "其他", "")]
work_other = [r for r in work_rows if r[0] not in ("海興段", "其他", "")]

def work_table(rows):
    return "".join(
        f"<tr><td class='cat'>{esc(r[0])}</td><td>{esc(r[1])}</td>"
        f"<td class='note'>{esc(r[2])}</td><td class='note'>{esc(r[3])}</td></tr>"
        for r in rows)

work_html = work_table(work_93h)
work_other_html = work_table(work_other)

done_cnt = len(items) - len(open_items)

page = f"""<!DOCTYPE html>
<html lang="zh-Hant"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>93H 進度儀表板</title>
<link rel="apple-touch-icon" sizes="180x180" href="icon-180.png">
<link rel="icon" type="image/png" sizes="32x32" href="favicon-32.png">
<meta name="theme-color" content="#1a3c6e">
<meta property="og:title" content="93H 海興段 進度儀表板">
<meta property="og:description" content="好瀚建設 93H 海興段：里程碑倒數、出工預核、到期警示、缺失改善">
<meta property="og:image" content="https://jammieaiwriter-jpg.github.io/highhand/93h/icon-512.png">
<style>
:root {{ --red:#c0392b; --orange:#d68910; --blue:#2471a3; --green:#1e8449; --gray:#707b7c; }}
* {{ box-sizing:border-box; }}
body {{ font-family:"Microsoft JhengHei","PingFang TC",sans-serif; margin:0; background:#f4f6f7; color:#212f3c; }}
header {{ background:#1a3c6e; color:#fff; padding:14px 20px; }}
header h1 {{ margin:0; font-size:1.3em; }}
header .sub {{ font-size:.85em; opacity:.85; margin-top:4px; }}
.wrap {{ max-width:1100px; margin:0 auto; padding:12px; }}
.kpis {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(140px,1fr)); gap:10px; margin:12px 0; }}
.kpi {{ background:#fff; border-radius:10px; padding:12px; text-align:center; box-shadow:0 1px 3px rgba(0,0,0,.1); }}
.kpi .num {{ font-size:1.9em; font-weight:700; }}
.kpi .lbl {{ font-size:.82em; color:#555; }}
.k-red .num {{ color:var(--red); }} .k-orange .num {{ color:var(--orange); }}
.k-blue .num {{ color:var(--blue); }} .k-green .num {{ color:var(--green); }} .k-star .num {{ color:#8e44ad; }}
.msbar {{ display:flex; gap:8px; overflow-x:auto; padding:4px 0 10px; }}
.ms {{ min-width:150px; background:#fff; border-radius:10px; padding:10px; border-top:4px solid var(--gray); box-shadow:0 1px 3px rgba(0,0,0,.1); flex:1; }}
.ms-hot {{ border-top-color:var(--orange); }} .ms-past {{ border-top-color:var(--red); }} .ms-ok {{ border-top-color:var(--green); }}
.ms-date {{ font-weight:700; font-size:.9em; }}
.ms-name {{ font-size:.85em; margin:4px 0; }}
.ms-cd {{ font-size:.8em; font-weight:700; color:var(--blue); }}
.ms-note {{ font-size:.72em; color:#777; margin-top:3px; }}
.sec {{ background:#fff; border-radius:10px; margin:12px 0; box-shadow:0 1px 3px rgba(0,0,0,.1); overflow:hidden; }}
.sec summary {{ cursor:pointer; padding:12px 14px; font-weight:700; font-size:1.02em; background:#eaf0f6; }}
.sec .cnt {{ background:#1a3c6e; color:#fff; border-radius:10px; padding:1px 9px; font-size:.8em; margin-left:8px; }}
table {{ width:100%; border-collapse:collapse; font-size:.88em; }}
th {{ background:#f2f4f4; text-align:left; padding:7px 9px; border-bottom:2px solid #d5dbdb; white-space:nowrap; }}
td {{ padding:7px 9px; border-bottom:1px solid #eaeded; vertical-align:top; }}
td.cat {{ white-space:nowrap; color:#555; }}
td.note {{ color:#666; font-size:.92em; }}
.badge {{ display:inline-block; border-radius:6px; padding:2px 8px; font-size:.85em; white-space:nowrap; color:#fff; }}
.b-red {{ background:var(--red); }} .b-orange {{ background:var(--orange); }}
.b-blue {{ background:var(--blue); }} .b-green {{ background:var(--green); }} .b-gray {{ background:var(--gray); }}
.permit summary {{ background:#fdf2d0; }}
footer {{ text-align:center; color:#888; font-size:.78em; padding:16px; }}
.ms-head {{ padding:10px 14px 4px; }}
.ms-title {{ font-weight:700; font-size:1.05em; margin-right:10px; }}
.ms-meta {{ color:#555; font-size:.85em; }}
.bar {{ height:8px; background:#e5e8ec; border-radius:4px; margin:8px 0 4px; }}
.bar-in {{ height:8px; background:#1e8449; border-radius:4px; }}
.ms-sub {{ color:#888; font-size:.78em; }}
.ms-risk {{ color:#c0392b; font-size:.85em; font-weight:700; margin-top:4px; }}
.subh {{ padding:8px 14px 2px; font-weight:700; font-size:.9em; color:#333; }}
.pct {{ float:right; color:#1e8449; font-weight:700; }}
.ms-date-tag {{ margin-left:10px; color:#1a3c6e; font-weight:700; }}
.ms-cd-tag {{ margin-left:8px; color:#2471a3; font-weight:700; }}
.kpi-row-label {{ font-weight:700; margin:10px 0 4px; font-size:.95em; }}
.kpi-row-label.sub {{ color:#777; font-size:.85em; }}
.kpis-sub .kpi {{ padding:6px; }}
.kpis-sub .kpi .num {{ font-size:1.3em; }}
.kpis-sub .kpi .lbl {{ font-size:.72em; }}
.tw {{ overflow-x:auto; -webkit-overflow-scrolling:touch; }}
.tw table {{ min-width:620px; }}
@media (max-width:640px) {{
  header h1 {{ font-size:1.05em; }}
  header .sub {{ font-size:.72em; }}
  .wrap {{ padding:8px; }}
  .kpis {{ grid-template-columns:repeat(3,1fr); gap:6px; }}
  .kpi {{ padding:8px 4px; }}
  .kpi .num {{ font-size:1.4em; }}
  .kpi .lbl {{ font-size:.68em; }}
  .ms {{ min-width:128px; padding:8px; }}
  table {{ font-size:.8em; }}
  th, td {{ padding:5px 6px; }}
  .sec summary {{ font-size:.92em; padding:10px 12px; }}
  .badge {{ font-size:.78em; }}
}}
@media print {{ .sec {{ break-inside:avoid; }} }}
</style></head><body>
<header><h1>93H 海興段 進度儀表板</h1>
<div class="sub">產出：{minguo(TODAY)}（{TODAY.isoformat()}）｜資料：{esc(SRC_LABEL)}｜距取得使照 <b>{license_day} 天</b></div>
</header>
<div class="wrap">
{kpi_html()}
<div class="msbar">{''.join(ms_html)}</div>
{'<details class="sec dispatch"><summary>👷 出工預核（今日 ' + str(len(dispatch_today)) + ' 組）<span class="cnt">' + str(len(dispatch_upcoming)) + '</span></summary><table><tr><th>日期</th><th>廠商/工班</th><th>預定工作</th><th>來源</th><th>實際出工</th><th>備註</th></tr>' + ''.join(f"<tr{(' style=background:#fbe3e0' if re.search(r'未出|未完成|未進|未派', str(r[4])) else ' style=background:#fff7e0') if parse_date(r[0])==TODAY else ''}><td>{esc(r[0])}</td><td><b>{esc(r[1])}</b></td><td>{esc(r[2])}</td><td class='note'>{esc(r[3])}</td><td>{esc(r[4]) or '—'}</td><td class='note'>{esc(r[5])}</td></tr>" for r in dispatch_upcoming) + '</table></details>' if dispatch_upcoming else ''}
{'<details class="sec" style="border-left:4px solid #c0392b"><summary>🔴 未完成／改期待重排<span class="cnt">' + str(len(dispatch_failed) + len(admin_failed)) + '</span></summary><table><tr><th>原定日</th><th>廠商/對象</th><th>事項</th><th>狀態</th><th>備註</th></tr>' + ''.join(f"<tr style=background:#fbe3e0><td>{esc(r[0])}</td><td><b>{esc(r[1])}</b></td><td>{esc(r[2])}</td><td>{esc(r[4])}</td><td class='note'>{esc(r[5])}</td></tr>" for r in dispatch_failed) + ''.join(f"<tr style=background:#fbe3e0><td>{esc(r[0])}</td><td><b>{esc(r[1])}</b></td><td>{esc(r[2])}</td><td>{esc(r[4])}</td><td class='note'>{esc(r[5])}</td></tr>" for r in admin_failed) + '</table></details>' if (dispatch_failed or admin_failed) else ''}
{'<details class="sec"><summary>🚚 進料追蹤<span class="cnt">' + str(len(mat_upcoming) + len(mat_unverified)) + '</span></summary><table><tr><th>日期</th><th>廠商</th><th>料項</th><th>來源</th><th>實際到料</th><th>備註</th></tr>' + ''.join(f"<tr{' style=background:#fff7e0' if parse_date(r[0])==TODAY else ''}><td>{esc(r[0]) or '待定'}</td><td><b>{esc(r[1])}</b></td><td>{esc(r[2])}</td><td class='note'>{esc(r[3])}</td><td>{esc(r[4]) or '—'}</td><td class='note'>{esc(r[5])}</td></tr>" for r in mat_upcoming) + ''.join(f"<tr><td>{esc(r[0])}</td><td><b>{esc(r[1])}</b></td><td>{esc(r[2])}</td><td class='note'>{esc(r[3])}</td><td>❓未記到料</td><td class='note'>{esc(r[5])}</td></tr>" for r in mat_unverified) + '</table></details>' if (mat_upcoming or mat_unverified) else ''}
{'<details class="sec"><summary>📑 使照檢附盤點（建照附款）<span class="cnt">' + str(len(rows_of("使照檢附", 5))) + '</span></summary><table><tr><th>檢附項目</th><th>附款</th><th>主辦</th><th>目前狀態</th><th>備註</th></tr>' + ''.join(f"<tr style=background:{'#fff7e0' if re.search('已取得|已交|已完成|✓', r[3]) else '#fbe3e0'}><td><b>{esc(r[0])}</b></td><td>{esc(r[1])}</td><td>{esc(r[2])}</td><td>{esc(r[3])}</td><td class='note'>{esc(r[4])}</td></tr>" for r in rows_of("使照檢附", 5)) + '</table></details>' if "使照檢附" in wb.sheetnames else ''}
{'<details class="sec"><summary>🏛️ 送審與變更（審核單位，時程可能拖）<span class="cnt">' + str(len(submissions)) + '</span></summary><table><tr><th>事項</th><th>受理/審核單位</th><th>送件日</th><th>預計核准</th><th>實際核准</th><th>備註</th></tr>' + ''.join(f"<tr><td><b>{esc(s[0])}</b></td><td>{esc(s[1])}</td><td>{esc(s[2]) or '—'}</td><td>{esc(s[3]) or '—'}</td><td>{esc(s[4]) or '—'}</td><td class='note'>{esc(s[6])}</td></tr>" for s in submissions) + '</table></details>' if submissions else ''}
{'<details class="sec"><summary>📅 行事曆與交辦<span class="cnt">' + str(len(admin_upcoming) + len(admin_unverified)) + '</span></summary><table><tr><th>日期</th><th>對象/窗口</th><th>事項</th><th>來源</th><th>完成</th><th>備註</th></tr>' + ''.join(f"<tr{' style=background:#fff7e0' if parse_date(r[0])==TODAY else ''}><td>{esc(r[0])}</td><td><b>{esc(r[1])}</b></td><td>{esc(r[2])}</td><td class='note'>{esc(r[3])}</td><td>{esc(r[4]) or '—'}</td><td class='note'>{esc(r[5])}</td></tr>" for r in admin_upcoming) + ''.join(f"<tr><td>{esc(r[0])}</td><td><b>{esc(r[1])}</b></td><td>{esc(r[2])}</td><td class='note'>{esc(r[3])}</td><td>❓未記結果</td><td class='note'>{esc(r[5])}</td></tr>" for r in admin_unverified) + '</table></details>' if (admin_upcoming or admin_unverified) else ''}
{'<details class="sec"><summary>🤝 廠商協調（送樣/圖說/工序/到場）<span class="cnt">' + str(len(rows_of('廠商協調', 6))) + '</span></summary><table><tr><th>日期</th><th>廠商</th><th>類型</th><th>事項</th><th>狀態</th><th>備註</th></tr>' + ''.join(f"<tr><td>{esc(r[0]) or '—'}</td><td><b>{esc(r[1])}</b></td><td>{esc(r[2])}</td><td>{esc(r[3])}</td><td>{esc(r[4])}</td><td class='note'>{esc(r[5])}</td></tr>" for r in rows_of('廠商協調', 6)) + '</table></details>' if '廠商協調' in wb.sheetnames else ''}
{'<details class="sec" style="border-left:4px solid #7030A0"><summary>⚖️ 待長官裁決<span class="cnt">' + str(len(decisions)) + '</span></summary><table><tr><th>提出日</th><th>事項</th><th>工地/廠商建議</th><th>備註</th></tr>' + ''.join(f"<tr><td>{esc(r[0])}</td><td><b>{esc(r[1])}</b></td><td class='note'>{esc(r[2])}</td><td class='note'>{esc(r[5])}</td></tr>" for r in decisions) + '</table></details>' if decisions else ''}
{'<details class="sec"><summary>🛠️ 缺失改善追蹤<span class="cnt">' + str(len(defects)) + '</span></summary><table><tr><th>發現日</th><th>缺失內容</th><th>位置</th><th>責任廠商</th><th>改善方式</th><th>狀態</th><th>備註</th></tr>' + ''.join(f"<tr><td>{esc(r[0])}</td><td>{esc(r[1])}</td><td>{esc(r[2])}</td><td><b>{esc(r[3])}</b></td><td class='note'>{esc(r[4])}</td><td>{esc(r[6])}</td><td class='note'>{esc(r[7])}</td></tr>" for r in defects) + '</table></details>' if defects else ''}
{'<details class="sec"><summary>⛏️ 粗工/打石統計（依費用歸屬）<span class="cnt">' + str(len(labor)) + '</span></summary><table><tr><th>費用歸屬</th><th>本月工數</th><th>本月金額</th><th>整場工數</th><th>整場金額</th></tr>' + ''.join(f"<tr><td><b>{esc(k)}</b></td><td>{v['m_n']:g}</td><td>{v['m_amt']:,.0f}</td><td>{v['t_n']:g}</td><td>{v['t_amt']:,.0f}</td></tr>" for k, v in sorted(labor_stats.items())) + '</table><table style="margin-top:8px"><tr><th>日期</th><th>工種</th><th>廠商</th><th>人數</th><th>工作內容</th><th>費用歸屬</th><th>備註</th></tr>' + ''.join(f"<tr><td>{esc(r[0])}</td><td>{esc(r[1])}</td><td>{esc(r[2])}</td><td>{esc(r[3])}</td><td class='note'>{esc(r[4])}</td><td><b>{esc(r[5])}</b></td><td class='note'>{esc(r[7])}</td></tr>" for r in labor[-15:]) + '</table></details>' if labor else ''}
{tree_html()}
<details class="sec"><summary>📋 發包與工作事項（93H）<span class="cnt">{len(work_93h)}</span></summary>
<table><tr><th>區/類</th><th>事項</th><th>現況</th><th>下一步</th></tr>{work_html}</table></details>
{'<details class="sec"><summary>🏘️ 其他案場（總安段/福智/新家波…）<span class="cnt">' + str(len(work_other)) + '</span></summary><table><tr><th>案場</th><th>事項</th><th>現況</th><th>下一步</th></tr>' + work_other_html + '</table></details>' if work_other else ''}
</div>
<footer>93H 進度儀表板｜每週三會後更新 93H進度盤點_*.xlsx 再重新產出｜⭐=使照前重點</footer>
<script>
(function() {{
  var loaded = Date.now();
  document.addEventListener("visibilitychange", function() {{
    if (document.visibilityState === "visible" && Date.now() - loaded > 5 * 60 * 1000) {{
      location.reload();
    }}
  }});
}})();
</script>
</body></html>"""

page = page.replace("<table", "<div class=\"tw\"><table").replace("</table>", "</table></div>")

# ---- 今日工地回報（即時）區塊：REPORT_API 有填才注入 ----
if REPORT_API:
    live_html = r"""
<details class="sec" open style="border-left:4px solid #1a5fb4"><summary>📥 今日工地回報（即時）<span class="cnt" id="lr-cnt">…</span></summary>
<div id="lr-body" style="padding:10px 14px;font-size:.9em;color:#666">載入中…</div></details>
<script>
(function(){
  var esc=function(s){ return String(s==null?"":s).replace(/</g,"&lt;"); };
  var fmtT=function(v){
    if(/^\d{1,2}:\d{2}/.test(String(v))) return String(v);
    var d=new Date(v);
    if(!isNaN(d)) return ("0"+d.getHours()).slice(-2)+":"+("0"+d.getMinutes()).slice(-2);
    return String(v);
  };
  var vkey=function(v){ return String(v||"").replace(/[(（].*$/,"").slice(0,2); };
  // 正式廠商名 ↔ 工地口語名 對照（比對用）
  var ALIAS={"品豪":"油漆","鴻成":"水電","尚和":"石材","阿賓":"石材","佶興":"木門","鈴鹿":"塗料","鈦翔":"地磚","冠維":"泥作","大港":"玻璃","勁揚":"磁磚","金豪":"磁磚","毅文":"粗工","一成":"水電","柯":"油漆"};
  var vkeys=function(v){ var k=vkey(v); var ks=[k]; if(ALIAS[k]) ks.push(ALIAS[k]); var m=String(v||"").match(/[(（]([^)）]+)/); if(m) ks.push(m[1].slice(0,2)); return ks; };
  Promise.all([
    fetch("__API__?today=1").then(r=>r.json()),
    fetch("report/plan.json?t="+Date.now(),{cache:"no-store"}).then(r=>r.json()).catch(function(){return null;})
  ]).then(function(res){
    var d=res[0], plan=res[1];
    var raw=(d.raw||[]), am=(d.am||[]), pm=(d.pm||[]), sv=(d.sv||[]);
    document.getElementById("lr-cnt").textContent=raw.length;
    if(!raw.length){ document.getElementById("lr-body").innerHTML="今日尚無工地回報"; return; }
    var h="";
    // 預計 vs 實際比對表（早報進來時，瀏覽器端即時計算，不等桌機）
    if(am.length && plan){
      var iso=new Date(); var isoStr=iso.getFullYear()+"-"+("0"+(iso.getMonth()+1)).slice(-2)+"-"+("0"+iso.getDate()).slice(-2);
      var day=(plan.days||[]).filter(function(x){return x.date===isoStr;})[0];
      var rows="", matchedAm={};
      if(day){
        day.items.forEach(function(it){
          var ks=vkeys(it.vendor), hit=null;
          am.forEach(function(a,i){
            if(hit) return;
            var hay=String(a[2]);   // 只比工班名稱欄，避免工作內容出現「水電」等字誤配
            for(var j=0;j<ks.length;j++){ if(ks[j] && hay.indexOf(ks[j])>=0){ hit=a; matchedAm[i]=true; return; } }
          });
          rows+="<tr><td>"+(hit?"✅":"🔴")+"</td><td><b>"+esc(it.vendor)+"</b></td><td class='note'>"+esc(it.work)+"</td><td>"+
                (hit? esc((hit[3]?hit[3]+"人：":"")+hit[4]) : "<span style='color:#c0392b'>未回報出工</span>")+"</td></tr>";
        });
        am.forEach(function(a,i){ if(!matchedAm[i]) rows+="<tr><td>➕</td><td><b>"+esc(a[2])+"</b></td><td class='note'>（預核外）</td><td>"+esc((a[3]?a[3]+"人：":"")+a[4])+"</td></tr>"; });
        h+="<div class='tw'><table style='margin:4px 0 10px'><tr><th></th><th>工班</th><th>預計工作</th><th>實際回報</th></tr>"+rows+"</table></div>";
        var abn=am.length&&am[0][6]? String(am[0][6]):"";
        if(abn) h+="<div style='color:#b45309;margin:0 0 8px'>⚠️ "+esc(abn).replace(/\n/g,"<br>")+"</div>";
      }
    }
    // 收工回報摘要（即時）：未答標紅、風險字標橘、明日派工置頂
    if(pm.length){
      var RISK=/未完成|順延|改期|延期|滲水|漏水|停工|損壞|無法|斷網|缺失|追加|未進場|未出/;
      var un=0, tomorrow="", items="";
      pm.forEach(function(q){
        var ans=String(q[4]||"").trim();
        var isUn=!ans||ans==="（未填）";
        if(isUn) un++;
        if(String(q[3]).indexOf("明日派工")>=0 && !isUn) tomorrow=ans;
        var col=isUn?"#c0392b":(RISK.test(ans)?"#b45309":"");
        items+="<div style='margin:5px 0;line-height:1.45;"+(col?"color:"+col+";":"")+"'><b>"+esc(q[2])+".</b> "+
               esc(String(q[3]).replace(/\s+/g," ").slice(0,32))+"<br>　@"+(isUn?"（未答）":esc(ans))+"</div>";
      });
      if(tomorrow) h+="<div style='margin:6px 0;padding:8px 10px;background:#eef4fb;border-radius:8px'>📌 <b>明日派工：</b>"+esc(tomorrow)+"</div>";
      h+="<details style='margin:6px 0'><summary style='cursor:pointer'><b>🌇 收工回報 "+pm.length+"題"+
         (un? "　<span style='color:#c0392b'>🔴"+un+"題未答</span>":"　✅全數作答")+"</b>（點開逐題）</summary>"+
         "<div style='padding:6px 2px'>"+items+"</div></details>";
    }
    // 會勘逐點（即時）
    if(sv.length){
      var svItems="";
      sv.forEach(function(s){ svItems+="<div style='margin:4px 0'><b>"+esc(s[4])+".</b> "+esc(s[5])+"｜"+esc(s[6])+(s[7]?"（照片"+esc(s[7])+"張）":"")+"</div>"; });
      h+="<details open style='margin:6px 0'><summary style='cursor:pointer'><b>📋 會勘紀錄 "+sv.length+"點</b>"+
         (sv[0][2]?"　巡視："+esc(sv[0][2]):"")+"</summary><div style='padding:6px 2px'>"+svItems+"</div></details>";
    }
    raw.slice().reverse().forEach(function(r){
      h+="<details style='margin:6px 0'><summary style='cursor:pointer'><b>"+fmtT(r[1])+"</b>　"+r[2]+
         "</summary><pre style='white-space:pre-wrap;background:#f6f7f9;padding:10px;border-radius:8px;font-size:.95em'>"+
         esc(r[3])+"</pre></details>";
    });
    document.getElementById("lr-body").innerHTML=h;
  }).catch(function(){ document.getElementById("lr-body").textContent="讀取失敗（稍後自動重試，或下拉重新整理）"; });
})();
</script>""".replace("__API__", REPORT_API)
    page = page.replace('<div class="msbar">', live_html + '\n<div class="msbar">', 1)

with open(OUT_HTML, "w", encoding="utf-8") as f:
    f.write(page)
print(f"OK -> {OUT_HTML}")

# ---- 工地回報頁用：未來7天預計出工 plan.json（隨 bat 一起推上 GitHub Pages）----
plan_days = []
for i in range(0, 8):
    d = TODAY + timedelta(days=i)
    items = [{"vendor": r[1], "work": r[2], "source": r[3], "note": r[5]}
             for r in dispatch_all if parse_date(r[0]) == d]
    if items:
        plan_days.append({"date": d.isoformat(), "roc": minguo(d), "items": items})
_plan_dir = os.environ.get("PLAN_DIR") or os.path.join(BASE, "highhand_repo", "93h", "report")
PLAN_PATH = os.path.join(_plan_dir, "plan.json")
if os.path.isdir(os.path.dirname(PLAN_PATH)):
    with open(PLAN_PATH, "w", encoding="utf-8") as f:
        json.dump({"generated": TODAY.isoformat(), "days": plan_days}, f, ensure_ascii=False, indent=1)
    print(f"OK -> {PLAN_PATH}")
print(f"統計：逾期{len(overdue)} 兩週內{len(soon)} 可施作{len(ready)} 施作中{len(doing)} 使照前未完{len(permit_open)} 排程{len(later)} 等前置{len(waiting)} 完成{done_cnt}/{len(items)}")
